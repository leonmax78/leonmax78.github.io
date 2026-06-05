from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DESKTOP = Path.home() / "Desktop"
DEFAULT_SETTING = DESKTOP / "0525" / "SETTING"
RAW_DIR = ROOT / "raw"

UNCATEGORIZED = "\u672a\u5206\u985e"
BEAST_SKILLS = ["\u8ff7\u60d1\u8853", "\u596a\u5fc3\u8853", "\u7e1b\u9748\u8853"]
BEAST_FALLBACK_CATEGORY = "\u672a\u5206\u985e"
KIND_ARTIFACT = "\u6cd5\u5668"
KIND_RECIPE = "\u914d\u65b9"


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "V" if value else ""
    return str(value).strip()


def uniq(values) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = clean(value)
        if not text or text in seen:
            continue
        seen.add(text)
        out.append(text)
    return out


def norm_name(value) -> str:
    return re.sub(r"[\s\[\]【】<>＜＞「」『』《》〈〉（）()◢◣◥◤▾▴▸◂▼▲◆◇■□▓★☆‧·．・]", "", clean(value))


def split_list(value) -> list[str]:
    text = clean(value)
    if not text:
        return []
    parts: list[str] = []
    buf: list[str] = []
    depth = 0
    opens = {"\u3010", "[", "(", "（", "<", "＜"}
    closes = {"\u3011", "]", ")", "）", ">", "＞"}
    separators = {"\u3001", ",", "\uff0c", ";", "\uff1b", "\n", "\r"}
    for char in text:
        if char in opens:
            depth += 1
        elif char in closes and depth > 0:
            depth -= 1
        if char in separators and depth == 0:
            part = "".join(buf).strip()
            if part:
                parts.append(part)
            buf = []
            continue
        buf.append(char)
    tail = "".join(buf).strip()
    if tail:
        parts.append(tail)
    return uniq(parts)


def find_one(pattern: str, label: str) -> Path:
    matches = [p for p in DESKTOP.glob(pattern) if not p.name.startswith("~$")]
    if not matches:
        raise FileNotFoundError(f"Cannot find {label} on Desktop with pattern: {pattern}")
    return max(matches, key=lambda p: p.stat().st_mtime)


def parse_ini_records(path: Path, section: str) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    cur: dict[str, str] | None = None
    try:
        text = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError:
        text = path.read_text(encoding="mbcs", errors="ignore")
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if (
            not line
            or line.startswith("//")
            or line.startswith("#include")
            or line.startswith("[DEFINE]")
            or line.startswith("#define")
        ):
            continue
        if line.startswith("[") and line.endswith("]"):
            if cur and cur.get("_section") == section:
                records.append(cur)
            cur = {"_section": line[1:-1]}
            continue
        if cur is not None and "=" in line:
            key, value = line.split("=", 1)
            cur[key.strip()] = value.strip()
    if cur and cur.get("_section") == section:
        records.append(cur)
    return records


def build_name_indexes(records: list[dict[str, str]]) -> tuple[dict[str, dict[str, str]], dict[str, list[str]]]:
    by_id: dict[str, dict[str, str]] = {}
    ids_by_name: dict[str, list[str]] = {}
    for record in records:
        row_id = clean(record.get("ID"))
        name = clean(record.get("Name"))
        if row_id:
            by_id[row_id] = record
        if name and row_id:
            ids_by_name.setdefault(name, []).append(row_id)
    return by_id, ids_by_name


def parse_collectbook(path: Path, item_by_id: dict[str, dict], monster_by_id: dict[str, dict]):
    collect_item_ids: set[str] = set()
    collect_monster_ids: set[str] = set()
    collect_index_by_id: dict[str, str] = {}
    for record in parse_ini_records(path, "LIST"):
        row_id = clean(record.get("ID"))
        if row_id:
            collect_item_ids.add(row_id)
            collect_index_by_id[row_id] = clean(record.get("Index"))
        for key, value in record.items():
            if key in ("_section", "ID", "Index", "Cost"):
                continue
            for number in re.findall(r"\d+", clean(value)):
                if number in monster_by_id:
                    collect_monster_ids.add(number)
    for row_id in list(collect_item_ids):
        if row_id not in item_by_id and row_id in monster_by_id:
            collect_monster_ids.add(row_id)
    return collect_item_ids, collect_monster_ids, collect_index_by_id


def load_locations(path: Path) -> dict[str, list[str]]:
    loc_by_name: dict[str, list[str]] = {}
    if not path.exists():
        return loc_by_name
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.reader(handle):
            if len(row) >= 2 and clean(row[0]):
                loc_by_name.setdefault(clean(row[0]), []).append(clean(row[1]))
    return {name: uniq(values) for name, values in loc_by_name.items()}


def parse_drop(value) -> list[tuple[str, float]]:
    nums = [x.strip() for x in str(value or "").split(",") if x.strip()]
    raw: list[tuple[str, float]] = []
    for i in range(2, len(nums) - 1, 2):
        item_id = nums[i]
        try:
            weight = float(nums[i + 1])
        except ValueError:
            weight = 0
        if item_id and item_id != "0" and weight > 0:
            raw.append((item_id, weight))
    total = sum(weight for _, weight in raw)
    return [(item_id, weight / total * 100 if total else 0) for item_id, weight in raw]


def build_drop_reverse(monsters: list[dict[str, str]], loc_by_name: dict[str, list[str]]):
    reverse: dict[str, list[dict]] = {}
    for monster in monsters:
        monster_id = clean(monster.get("ID"))
        monster_name = clean(monster.get("Name"))
        monster_level = clean(monster.get("Level"))
        locations = loc_by_name.get(monster_name, [])
        if monster_name == "\u2593\u554f\u9802\u4ed9\u9f8d\u2593" and monster_level == "1200":
            locations = uniq(locations + ["\u7d55\u4e03\u5bf6\u4ed9\u5883"])
        for item_id, rate in parse_drop(monster.get("DropItem")):
            reverse.setdefault(item_id, []).append(
                {
                    "monsterId": monster_id,
                    "monster": monster_name,
                    "level": monster_level,
                    "rate": round(rate, 6),
                    "locations": locations,
                }
            )
    for rows in reverse.values():
        rows.sort(key=lambda row: (-row["rate"], row["monster"], row["monsterId"]))
    return reverse


def load_shop_maps(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    maps = {"weapon": {}, "artifact": {}, "recipe": {}}
    columns = {"weapon": (1, 2, 3), "artifact": (4, 5, 6), "recipe": (7, 8, 9)}
    for row in ws.iter_rows(min_row=4, values_only=True):
        for key, (name_col, score_col, location_col) in columns.items():
            name = clean(row[name_col - 1])
            if not name:
                continue
            maps[key][name] = {
                "score": clean(row[score_col - 1]),
                "shops": split_list(row[location_col - 1]),
            }
    return maps


def load_weapon_categories(wb):
    category_by_name: dict[str, str] = {}
    category_order: list[str] = []
    for ws in wb.worksheets[5:21]:
        category_order.append(ws.title)
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = clean(row[0])
            if name:
                category_by_name.setdefault(name, ws.title)
    return category_by_name, category_order


def choose_id(name: str, ids_by_name: dict[str, list[str]], collect_ids: set[str]) -> tuple[str, list[str]]:
    ids = ids_by_name.get(name, [])
    best = next((row_id for row_id in ids if row_id in collect_ids), ids[0] if ids else "")
    return best, ids


def rows_from_collect_sheet(
    ws,
    kind: str,
    shop_key: str,
    shop_maps,
    category_by_name,
    item_ids_by_name,
    collect_item_ids,
    collect_index_by_id,
    drop_reverse,
):
    rows = []
    current_category = UNCATEGORIZED
    for source_row in ws.iter_rows(min_row=2, values_only=True):
        name = clean(source_row[0])
        if not name:
            continue
        score = clean(source_row[1])
        task_flag = clean(source_row[2])
        shop_flag = clean(source_row[3])
        sources = uniq(source_row[4:])
        shops = shop_maps.get(shop_key, {}).get(name, {}).get("shops", [])
        if shop_flag and not shops:
            shops = sources[:]
        task_names = split_list(task_flag) if task_flag and task_flag.upper() != "V" else []
        item_id, all_item_ids = choose_id(name, item_ids_by_name, collect_item_ids)
        reverse_drops = drop_reverse.get(item_id, []) if item_id else []
        if kind == "weapon":
            category = category_by_name.get(name, current_category)
            if category != UNCATEGORIZED:
                current_category = category
        elif kind == "artifact":
            category = KIND_ARTIFACT
        else:
            category = KIND_RECIPE
        rows.append(
            {
                "name": name,
                "itemId": item_id,
                "allItemIds": all_item_ids,
                "collectIndex": collect_index_by_id.get(item_id, ""),
                "kind": kind,
                "category": category,
                "score": score,
                "taskFlag": bool(task_flag),
                "taskNames": task_names,
                "shopFlag": bool(shop_flag),
                "shops": shops,
                "excelSources": sources,
                "reverseDrops": reverse_drops,
                "reverseDropCount": len(reverse_drops),
                "searchText": " ".join(
                    [
                        name,
                        item_id,
                        category,
                        score,
                        " ".join(task_names),
                        " ".join(shops),
                        " ".join(sources),
                        " ".join(row["monster"] for row in reverse_drops[:100]),
                    ]
                ),
            }
        )
    return rows


def beast_rows(ws, monster_ids_by_name, collect_monster_ids, collect_index_by_id, loc_by_name):
    rows = []
    current_group = BEAST_FALLBACK_CATEGORY
    for source_row in ws.iter_rows(min_row=2, values_only=True):
        if clean(source_row[0]):
            current_group = clean(source_row[0])
        name = clean(source_row[1])
        if not name:
            continue
        monster_id, all_monster_ids = choose_id(name, monster_ids_by_name, collect_monster_ids)
        locations = split_list(source_row[7]) if clean(source_row[7]) else loc_by_name.get(name, [])
        skills = [label for label, cell in zip(BEAST_SKILLS, source_row[4:7]) if clean(cell)]
        rows.append(
            {
                "name": name,
                "monsterId": monster_id,
                "allMonsterIds": all_monster_ids,
                "collectIndex": collect_index_by_id.get(monster_id, ""),
                "kind": "beast",
                "category": current_group,
                "score": clean(source_row[2]),
                "strength": clean(source_row[3]),
                "skills": skills,
                "locations": locations,
                "searchText": " ".join([name, monster_id, current_group, clean(source_row[2]), clean(source_row[3]), " ".join(locations)]),
            }
        )
    return rows


def build(args) -> dict:
    collect_workbook = Path(args.collect_workbook) if args.collect_workbook else find_one("*2025_8_28*.xlsx", "collect workbook")
    shop_workbook = Path(args.shop_workbook) if args.shop_workbook else find_one("*2025_1_21*.xlsx", "shop workbook")
    setting_dir = Path(args.setting_dir)
    item_ini = RAW_DIR / "ITEM.INI" if (RAW_DIR / "ITEM.INI").exists() else setting_dir / "ITEM.INI"
    monster_ini = (
        RAW_DIR / "MONSTER_C_MERGED.INI"
        if (RAW_DIR / "MONSTER_C_MERGED.INI").exists()
        else setting_dir / "MONSTER_C.INI"
    )
    collectbook_ini = setting_dir / "COLLECTBOOKITEM.INI"

    items = parse_ini_records(item_ini, "ITEM")
    monsters = parse_ini_records(monster_ini, "NPC")
    item_by_id, item_ids_by_name = build_name_indexes(items)
    monster_by_id, monster_ids_by_name = build_name_indexes(monsters)
    collect_item_ids, collect_monster_ids, collect_index_by_id = parse_collectbook(
        collectbook_ini, item_by_id, monster_by_id
    )
    loc_by_name = load_locations(Path(args.locations_csv))
    drop_reverse = build_drop_reverse(monsters, loc_by_name)
    shop_maps = load_shop_maps(shop_workbook)

    wb = load_workbook(collect_workbook, read_only=True, data_only=True)
    category_by_name, category_order = load_weapon_categories(wb)
    data = {
        "meta": {
            "sourceCollectWorkbook": collect_workbook.name,
            "sourceShopWorkbook": shop_workbook.name,
            "sourceCollectBook": str(collectbook_ini),
            "sourceItemIni": str(item_ini),
            "sourceMonsterIni": str(monster_ini),
            "itemRecords": len(items),
            "monsterRecords": len(monsters),
            "weaponCategoryOrder": category_order,
            "orderPolicy": "Rows are emitted in source workbook order; weapon category order follows the workbook tab order; new official rows appended at source end remain appended in JSON.",
        },
        "weapon": rows_from_collect_sheet(
            wb.worksheets[0],
            "weapon",
            "weapon",
            shop_maps,
            category_by_name,
            item_ids_by_name,
            collect_item_ids,
            collect_index_by_id,
            drop_reverse,
        ),
        "artifact": rows_from_collect_sheet(
            wb.worksheets[1],
            "artifact",
            "artifact",
            shop_maps,
            category_by_name,
            item_ids_by_name,
            collect_item_ids,
            collect_index_by_id,
            drop_reverse,
        ),
        "recipe": rows_from_collect_sheet(
            wb.worksheets[2],
            "recipe",
            "recipe",
            shop_maps,
            category_by_name,
            item_ids_by_name,
            collect_item_ids,
            collect_index_by_id,
            drop_reverse,
        ),
        "beast": beast_rows(wb.worksheets[3], monster_ids_by_name, collect_monster_ids, collect_index_by_id, loc_by_name),
    }
    return data


def main() -> int:
    parser = argparse.ArgumentParser(description="Build data/collectbook_sources.json from Excel and INI sources.")
    parser.add_argument("--collect-workbook", default="")
    parser.add_argument("--shop-workbook", default="")
    parser.add_argument("--setting-dir", default=str(DEFAULT_SETTING))
    parser.add_argument("--locations-csv", default=str(RAW_DIR / "\u4e00\u822c\u602a\u7269\u4f4d\u7f6e.csv"))
    parser.add_argument("--out", default=str(ROOT / "data" / "collectbook_sources.json"))
    args = parser.parse_args()

    data = build(args)
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {out_path}")
    print({key: len(data[key]) for key in ("weapon", "artifact", "recipe", "beast")})
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
